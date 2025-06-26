import requests
import pandas as pd
import time
from datetime import datetime
from typing import List
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill
from tkinter import Tk, filedialog, Text, Button, Label, END
import urllib3
from openpyxl.chart import PieChart, Reference
from openpyxl.utils import get_column_letter


urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# --- CVE Class ---
class CVEDetails:
    def __init__(self, cve_id, description, cvss_score, severity, published_date, cwe_id):
        self.cve_id = cve_id
        self.description = description
        self.cvss_score = cvss_score
        self.severity = severity
        self.published_date = published_date
        self.cwe_id = cwe_id

# --- NVD API Client ---
class NVDApiClient:
    def __init__(self, api_key: str):
        self.api_key = api_key
        self.base_url = "https://services.nvd.nist.gov/rest/json/cves/2.0"
        self.headers = {
            "X-Api-Key": api_key,
            "User-Agent": "Mozilla/5.0 (compatible; VulnerabilityScanner/1.0)"
        }
        self.request_delay = 6

    def get_cve_details(self, keyword: str, max_results: int = 1) -> List[CVEDetails]:
        try:
            params = {
                "keywordSearch": keyword,
                "resultsPerPage": max_results
            }

            time.sleep(self.request_delay)
            response = requests.get(self.base_url, headers=self.headers, params=params, timeout=30,verify=False)

            if response.status_code == 403:
                return []

            response.raise_for_status()
            data = response.json()

            cve_list = []
            for vuln in data.get("vulnerabilities", []):
                cve = vuln.get("cve", {})

                cvss_score = 0.0
                severity = "UNKNOWN"
                cwe_id = "N/A" # Initialize cwe_id

                metrics = cve.get("metrics", {})
                if metrics.get("cvssMetricV31"):
                    cvss_data = metrics["cvssMetricV31"][0].get("cvssData", {})
                    cvss_score = float(cvss_data.get("baseScore", 0.0))
                    severity = cvss_data.get("baseSeverity", "UNKNOWN")
                elif metrics.get("cvssMetricV30"):
                    cvss_data = metrics["cvssMetricV30"][0].get("cvssData", {})
                    cvss_score = float(cvss_data.get("baseScore", 0.0))
                    severity = cvss_data.get("baseSeverity", "UNKNOWN")
                elif metrics.get("cvssMetricV2"):
                    cvss_data = metrics["cvssMetricV2"][0].get("cvssData", {})
                    cvss_score = float(cvss_data.get("baseScore", 0.0))
                    severity = "HIGH" if cvss_score >= 7 else "MEDIUM" if cvss_score >= 4 else "LOW"
                    
                # Extract CWE
                weaknesses = cve.get("weaknesses", [])
                if weaknesses:
                    cwe_id = ", ".join([weakness.get("description", [{}])[0].get("value", "N/A") for weakness in weaknesses])


                description = "No description available"
                for desc in cve.get("descriptions", []):
                    if desc.get("lang") == "en":
                        description = desc.get("value", "No description available")
                        break

                published_date = cve.get("published", "Unknown")


                cve_details = CVEDetails(
                    cve_id=cve.get("id", "Unknown"),
                    description=description,
                    cvss_score=cvss_score,
                    severity=severity,
                    published_date=published_date,
                    cwe_id=cwe_id
                )
                cve_list.append(cve_details)

            return cve_list

        except Exception as e:
            print(e) # Print the exception to help debugging
            return []

# --- EPSS Fetcher ---
def get_epss_details(cve_id):
    api_url = f"https://api.first.org/data/v1/epss?cve={cve_id}"
    try:
        response = requests.get(api_url, timeout=10,verify=False)
        response.raise_for_status()
        data = response.json()
        item = data.get('data', [{}])[0]
        return {
            'cve': item.get('cve', cve_id),
            'epss': item.get('epss', 'N/A'),
            'percentile': item.get('percentile', 'N/A'),
            'date': item.get('date', 'N/A')
        }
    except Exception:
        return {
            'cve': cve_id,
            'epss': 'Error',
            'percentile': 'Error',
            'date': 'Error'
        }

def analyze_cwe_ids(cve_details_list: List[CVEDetails]) -> dict:
    """
    Analyzes the CWE IDs from a list of CVEDetails objects and returns a dictionary
    containing the counts of each CWE ID.

    Args:
        cve_details_list (List[CVEDetails]): A list of CVEDetails objects.

    Returns:
        dict: A dictionary where keys are CWE IDs and values are their counts.
    """
    cwe_counts = {}
    for cve_details in cve_details_list:
        cwe_ids = cve_details.cwe_id.split(", ")  # Split multiple CWEs
        for cwe_id in cwe_ids:
             if cwe_id != "N/A": #skip N/A, not helpful for graphing.
                cwe_counts[cwe_id] = cwe_counts.get(cwe_id, 0) + 1
    return cwe_counts


def create_excel_sheet_with_pie_chart(cwe_counts: dict, workbook: Workbook, sheet_name: str = "CWE Analysis"):
    """
    Creates an Excel file with a pie chart representing CWE ID distribution on a new sheet.

    Args:
        cwe_counts (dict): A dictionary where keys are CWE IDs and values are their counts.
        workbook (Workbook): The excel workbook where the new sheet will be created
        sheet_name (str): The name of the new sheet for CWE analysis (default: "CWE Analysis").
    """
    sheet = workbook.create_sheet(sheet_name)  # Create the new sheet

    # Prepare data for the chart
    data = [["CWE ID", "Count"]]  # Headers
    for cwe_id, count in cwe_counts.items():
        data.append([cwe_id, count])

    # Write data to the sheet
    for row_idx, row in enumerate(data, 1):  # Start from row 1
        for col_idx, value in enumerate(row, 1):  # Start from column 1
            col_letter = get_column_letter(col_idx)
            sheet[f"{col_letter}{row_idx}"] = value

    # Create Pie Chart
    pie_chart = PieChart()
    pie_chart.title = "CWE ID Distribution"

    labels = Reference(sheet, min_col=1, min_row=2, max_row=len(cwe_counts) + 1)  # CWE ID labels
    values = Reference(sheet, min_col=2, min_row=2, max_row=len(cwe_counts) + 1)  # CWE ID counts

    pie_chart.add_data(values, titles_from_data=False) #values from where data will be populated
    pie_chart.set_categories(labels) #setting labels

    # Add the chart to the sheet
    sheet.add_chart(pie_chart, "E2")  # Place the chart in cell E2

def find_most_common_cwe(cwe_counts: dict) -> str:
    """
    Finds the most common CWE ID from a dictionary of CWE ID counts.

    Args:
        cwe_counts (dict): A dictionary where keys are CWE IDs and values are their counts.

    Returns:
        str: The most common CWE ID.
    """
    most_common_cwe = None
    max_count = 0

    for cwe_id, count in cwe_counts.items():
        if count > max_count:
            most_common_cwe = cwe_id
            max_count = count

    return most_common_cwe if most_common_cwe else "No CWEs found"

# === GUI Setup ===
def browse_file():
    filepath = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
    if filepath:
        label_file.config(text=filepath)
        run_button.config(state="normal")

def run_script():
    file_path = label_file.cget("text")
    output_box.insert(END, f"Loaded file: {file_path}\n")
    input_df = pd.read_excel(file_path)
    keywords = input_df['keyword'].dropna().tolist()

    api_key = "b0c2943a-8702-4560-b2b3-08ec65345d7b"
    client = NVDApiClient(api_key)
    kev_url = "https://www.cisa.gov/sites/default/files/csv/known_exploited_vulnerabilities.csv"
    kev_file = "known_exploited_vulnerabilities.csv"

    response = requests.get(kev_url, timeout=10,verify=False)
    with open(kev_file, "wb") as f:
        f.write(response.content)
    kev_df = pd.read_csv(kev_file)
    kev_df['cveID'] = kev_df['cveID'].str.lower()

    final_results = []
    for keyword in keywords:
        cves = client.get_cve_details(keyword)
        output_box.insert(END, f" Searching for keyword: {keyword}\n")
        output_box.see(END)
        for cve in cves:
            epss = get_epss_details(cve.cve_id)
            row = {
                'CVE ID': cve.cve_id,
                'Description': cve.description,
                'CVSS Score': cve.cvss_score,
                'Severity': cve.severity,
                'Published Date': cve.published_date,
                'CWE ID': cve.cwe_id,
                'EPSS Score': epss['epss'],
                'Percentile': epss['percentile'],
                'EPSS Date': epss['date']
            }
            final_results.append(row)
            output_box.insert(END, f"Processed {cve.cve_id}\n")
            output_box.see(END)

    df = pd.DataFrame(final_results)
    df['CVE ID'] = df['CVE ID'].str.lower()
    df['In_KEV'] = df['CVE ID'].isin(kev_df['cveID']).map({True: 'Yes', False: 'No'})

    filename = f"output_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
    df.to_excel(filename, index=False)

    wb = load_workbook(filename)
    source_ws = wb['Sheet1']
    ws_intro = wb.create_sheet(title='Introduction')

    ws_intro["A1"] = "This Python script automates the process of vulnerability analysis by leveraging the NIST NVD API and the FIRST EPSS API to gather detailed information about CVEs (Common Vulnerabilities and Exposures). The script takes an Excel file as input, reads a list of keywords from it, and queries the NVD API for corresponding CVE details. It then fetches EPSS (Exploit Prediction Scoring System) data for each CVE from the FIRST API. The script consolidates the gathered information, including descriptions, CVSS scores, severity levels, published dates, CWE ID, EPSS scores, and whether the CVE is listed in the CISA's Known Exploited Vulnerabilities (KEV) catalog. The results are written to a new Excel file. The script also includes a basic GUI built with Tkinter, allowing users to select an input Excel file and trigger the analysis. The generated Excel file includes color-coded formatting in the 'Formatted_Data' sheet, designed to highlight different priority levels based on CVSS and EPSS scores, making it easy to identify critical vulnerabilities. The script also create an introduction sheet with the purpose of the data table. Finally, it reorganizes the sheets to show them in proper order."

# 2. Starting at cell C5, create the table
    table_data = [
        ("Fill Color", "Priority", "Range"),
        ("Red", "P1", "In Kev"),
        ("Orange", "P2", "CVSS > 8 & EPSS > 0.7"),
        ("Yellow", "P3", "CVSS < 8 & > 4  // EPSS < 0.8 & > 0.4"),
        ("Green", "P4", "Low Priority"),
    ]

    # Fill colors for each row
    color_map = {
        "Red": "FFC7CE",
        "Orange": "FFEB9C",
        "Yellow": "FFFACD",
        "Green": "C6EFCE",
    }

    start_row = 5
    start_col = 3  # Column C

    # Write table data and apply fill colors
    for i, row in enumerate(table_data):
        for j, value in enumerate(row):
            cell = ws_intro.cell(row=start_row + i, column=start_col + j, value=value)

            # Apply fill to the first column (color cell), skipping header row
            if i != 0 and j == 0:
                fill = PatternFill(start_color=color_map[value], end_color=color_map[value], fill_type="solid")
                cell.fill = fill

    new_ws = wb.create_sheet(title='Formatted_Data')
    for row in source_ws.iter_rows():
            for cell in row:
                new_ws.cell(row=cell.row, column=cell.column, value=cell.value)
    source_ws.title = "Input Data"
    ws = wb['Formatted_Data']
    # === Define fills ===
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")       # CRITICAL / High EPSS
    orange_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")    # HIGH / Medium EPSS
    yellow_fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")    # MEDIUM
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    # === Apply coloring ===
    header_font = Font(bold=True, color="FFFFFF")  # White font color
    header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")  # Black background
    for col in ws.iter_cols(min_row=1, max_row=1, min_col=1, max_col=ws.max_column):
        col[0].font = header_font
        col[0].fill = header_fill

    # Get the list of CVEDetails objects after processing CVEs
    cve_details_list = [
        CVEDetails(
            cve_id=row[0].value,
            description=row[1].value,
            cvss_score=row[2].value,
            severity=row[3].value,
            published_date=row[4].value,
            cwe_id=row[5].value,
        )
        for row in new_ws.iter_rows(min_row=2, min_col=1, max_col=6)
    ]

    # Analyze CWE IDs and create the pie chart
    cwe_counts = analyze_cwe_ids(cve_details_list)
    create_excel_sheet_with_pie_chart(cwe_counts, wb)
    most_common_cwe = find_most_common_cwe(cwe_counts)
    output_box.insert(END, f"\nThe most common CWE ID in the report is: {most_common_cwe}\n")
    output_box.see(END)

    for row in ws.iter_rows(min_row=2, min_col=1, max_col=10):
        # Initialize variables
        fill = green_fill  # Default fill color
        try:
            in_kev_cell = row[9].value  # Column J is the 10th column (index 9)
            CVSS_cell = float(row[2].value)

            # Get EPSS value and handle 'N/A' or 'Error'
            epss_value = row[6].value
            if isinstance(epss_value, str) and (epss_value.lower() in ['n/a', 'error']):
                EPSS_cell = 0.0  # Default value if EPSS is not available
            else:
                EPSS_cell = float(epss_value)

        except ValueError as e:
            print(f"Error processing row: {row}.  Error: {e}")
            EPSS_cell = 0.0  # Set EPSS_cell to 0.0 on error
            # Continue to the coloring logic with a default EPSS value
            
        #Apply logic color
        if in_kev_cell == 'Yes':
            fill = red_fill
        elif CVSS_cell >= 8 and EPSS_cell > 0.7:
            fill = orange_fill
        elif 4 <= CVSS_cell < 8 and 0.4 < EPSS_cell <= 0.7:
            fill = yellow_fill
        elif 6 <= CVSS_cell < 10  and EPSS_cell < 0.4:
            fill = yellow_fill

        # Apply the fill to all cells in the row
        for cell in row:
            cell.fill = fill

    new_order = ["Introduction", "Input Data", "Formatted_Data", "CWE Analysis"]

# Reorder the sheets
    wb._sheets = [wb[sheet_name] for sheet_name in new_order]
    wb.save(filename)
    output_box.insert(END, f"\n Excel saved: {filename}\n")
    output_box.see(END)

# === Build UI ===
root = Tk()
root.title("Final CVE Analysis")
root.geometry("600x500")

label_intro = Label(root, text="Upload Final Report", font=("Arial", 12))
label_intro.pack(pady=10)

label_file = Label(root, text="No file selected", bg="white", width=60)
label_file.pack()

browse_button = Button(root, text=" Browse", command=browse_file)
browse_button.pack(pady=5)

run_button = Button(root, text="Run Analysis", state="disabled", command=run_script)
run_button.pack(pady=10)

output_box = Text(root, height=20, width=70)
output_box.pack()

root.mainloop()